
# -*- coding: utf-8 -*-
# @Time    : 2020/6/30 0030 下午 11:06
# @Author  : dingqin


from openpyxl import load_workbook

def read_data(file_name, sheet_name):
    wb = load_workbook(file_name)
    sheet = wb[sheet_name]
    all_list = []
    for i in range(sheet.max_row-1):
        list_a = []
        for j in range(1,sheet.max_column-1):  # 1 2 3 4 5 6 7
            value = sheet.cell(row=i + 2, column=j).value
            list_a.append(value)
        all_list.append(list_a)
    return all_list

def write_data(file_name,sheet_name,row,column,value):       #此函数是写入结果到excel中
    wb= load_workbook(file_name)
    sheet = wb[sheet_name]
    # 行是跟着id走，行数=id+1
    sheet.cell(row=row, column=column).value =value
    # 保存
    wb.save('test_case.xlsx')



if __name__ == '__main__':
    print(read_data('test_case.xlsx', 'recharge'))
